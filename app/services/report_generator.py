"""
report_generator.py — Exporta análisis a Excel y PDF.
"""
import io
from datetime import datetime
from pathlib import Path
from app.utils.logger import get_logger

logger = get_logger(__name__)


def generate_excel(doc_data: dict, analysis: dict) -> bytes:
    """Genera un .xlsx con metadata + análisis IA. Retorna bytes."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws1 = wb.active
    ws1.title = "Resumen"
    _style_header(ws1, "A1", "DocuFlow — Reporte de Análisis")

    rows = [
        ("Archivo", doc_data.get("filename", "")),
        ("Tipo", doc_data.get("file_type", "").upper()),
        ("Páginas", doc_data.get("page_count", "—")),
        ("Procesado", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Tipo de documento", analysis.get("tipo_documento", "")),
        ("Score de confianza", f"{analysis.get('score_confianza', 0):.0%}"),
        ("", ""),
        ("Resumen", analysis.get("resumen", "")),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws1[f"A{i}"] = k
        ws1[f"B{i}"] = str(v)
        if k:
            ws1[f"A{i}"].font = Font(bold=True)
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 60

    # --- Hoja 2: Entidades ---
    ws2 = wb.create_sheet("Entidades")
    _style_header(ws2, "A1", "Entidades detectadas")
    entidades = analysis.get("entidades", {})
    row = 3
    for categoria, items in entidades.items():
        ws2[f"A{row}"] = categoria.capitalize()
        ws2[f"A{row}"].font = Font(bold=True)
        for item in items:
            row += 1
            ws2[f"B{row}"] = item
        row += 1
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    # --- Hoja 3: Campos clave ---
    ws3 = wb.create_sheet("Campos clave")
    _style_header(ws3, "A1", "Campos clave extraídos")
    campos = analysis.get("campos_clave", {})
    ws3["A3"] = "Campo"
    ws3["B3"] = "Valor"
    ws3["A3"].font = ws3["B3"].font = Font(bold=True)
    for i, (k, v) in enumerate(campos.items(), start=4):
        ws3[f"A{i}"] = k
        ws3[f"B{i}"] = str(v)
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_pdf(doc_data: dict, analysis: dict) -> bytes:
    """Genera un PDF simple con reportlab. Retorna bytes."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2.5*cm, rightMargin=2.5*cm)
    styles = getSampleStyleSheet()
    purple = colors.HexColor("#534AB7")

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=purple, fontSize=20, spaceAfter=6)
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                               textColor=purple, fontSize=13, spaceBefore=14, spaceAfter=4)
    body_style = styles["BodyText"]
    body_style.fontSize = 10
    body_style.leading = 14

    story = []
    story.append(Paragraph("DocuFlow — Reporte de Análisis", title_style))
    story.append(Paragraph(
        f"Archivo: <b>{doc_data.get('filename', '')}</b> · "
        f"Tipo: <b>{analysis.get('tipo_documento', '').capitalize()}</b> · "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        body_style
    ))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Resumen ejecutivo", h2_style))
    story.append(Paragraph(analysis.get("resumen", "Sin resumen disponible."), body_style))
    story.append(Spacer(1, 0.3*cm))

    # Entidades como tabla
    entidades = analysis.get("entidades", {})
    ent_rows = [["Categoría", "Valor"]]
    for cat, items in entidades.items():
        for item in items:
            ent_rows.append([cat.capitalize(), item])

    if len(ent_rows) > 1:
        story.append(Paragraph("Entidades detectadas", h2_style))
        t = Table(ent_rows, colWidths=[5*cm, 11*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), purple),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEEDFE")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#AFA9EC")),
            ("PADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)

    # Anomalías
    anomalias = analysis.get("anomalias", [])
    if anomalias:
        story.append(Paragraph("Anomalías detectadas", h2_style))
        for a in anomalias:
            story.append(Paragraph(f"• {a}", body_style))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _style_header(ws, cell: str, title: str):
    from openpyxl.styles import Font, PatternFill
    ws[cell] = title
    ws[cell].font = Font(bold=True, size=13, color="FFFFFF")
    ws[cell].fill = PatternFill("solid", fgColor="534AB7")
